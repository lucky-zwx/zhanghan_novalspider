import xml.dom.minidom as xmldom

import openpyxl
import requests
from bs4 import BeautifulSoup
from fontTools.ttLib import TTFont
import time

# python的字典用于将英文转换为数字
number_dict = {
    "period": ".",
    "zero": "0",
    "one": "1",
    "two": "2",
    "three": "3",
    "four": "4",
    "five": "5",
    "six": "6",
    "seven": "7",
    "eight": "8",
    "nine": "9"
}


def Getinfo(url, class_nov, all_num):
    # 将程序暂停0.5秒，防止爬取过快导致的拒绝服务
    time.sleep(0.5)
    # 模拟浏览器请求
    re_get = requests.get(url=url, headers={
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:60.0) Gecko/20100101 Firefox/60.0'})
    # 文本转码
    re_get.encoding = 'utf-8'
    # 判断该书的页面是否有效
    if "网址已失效" in re_get.text:
        print('该书已下架！')
        return 0
    # 代码美化,进行下一步的前端代码元素定位
    be = BeautifulSoup(re_get.text.replace('&#', ''), 'html.parser')
    # 定位，获取ｕｒｌ
    html_data = be.find('div', class_='book-info').find_all('span')[-1].get('class')
    # 获取小说的详细页面ｕｒｌ
    info = be.find('div', class_='book-info').find_all('span', class_=html_data)
    # 下载字体库文件
    Savewoff("https://qidian.gtimg.com/qd_anti_spider/" + html_data[0] + ".woff")

    # 获取小说标题
    noveltitle = str(be.find('div', class_='book-info').h1.em.text)

    # 获取小说作者
    novelwriter = str(be.find('a', class_='writer').text)

    # 获取小说字数
    orginnum = str(info[0].text)
    numlist = orginnum.split(';')
    novelnum = comnum(numlist=numlist)

    # 获取小说总推荐数
    orginnum = str(info[1].text)
    numlist = orginnum.split(';')
    novellike_all = comnum(numlist=numlist)

    # 获取小说的周推荐数
    orginnum = str(info[2].text)
    numlist = orginnum.split(';')
    novellike_week = comnum(numlist=numlist)

    # print(noveltitle, novelwriter, novelnum, novellike_all, novellike_week, class_nov, all_num)
    save_excle(noveltitle, novelwriter, novelnum, novellike_all, novellike_week, class_nov, all_num)
    return 1


def Savewoff(woff_url):
    # 下载字体库
    try:
        font_content = requests.get(url=woff_url, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:60.0) Gecko/20100101 Firefox/60.0'}).content
        with open('qidian.woff', 'wb') as f:
            f.write(font_content)
        font1 = TTFont('qidian.woff')
        font1.saveXML('qidian.xml')
    except BaseException:
        print("爬取过快")
        return


def comnum(numlist):
    #初始化字符串类型的number
    number = ''
    #将ｗｏｆｆ文件转化为ｘｍｌ文件，以获得相对应的编码信息
    getxml = xmldom.parse('qidian.xml')
    #获取ｘｍｌ文件的操作对象
    elementobj = getxml.documentElement
    #获取ｘｍｌ中ｍａｐ标签对象
    subElementObj = elementobj.getElementsByTagName("map")
    #遍历编码
    for num in numlist:
        #遍历ｘｍｌ中的编码
        for ele in subElementObj:
            #如果对应的编码相同，则获得编码对应的数字
            if num != '' and str(hex(int(num))) == str(ele.getAttribute('code')):
                number += number_dict[ele.getAttribute('name')]
    #返回解码出的数字
    return number


def Getnovellist():
    # 先设置excle的写行数,应为第一行为标题,从第二行开始写
    all_num = 2
    # 分类页的ｕｒｌ
    list_ = {
        "https://www.qidian.com/rank/readIndex?style=1&chn=21&page=": "玄幻",
        "https://www.qidian.com/rank/readIndex?style=1&chn=1&page=": "奇幻",
        "https://www.qidian.com/rank/readIndex?style=1&chn=2&page=": "武侠",
        "https://www.qidian.com/rank/readIndex?style=1&chn=22&page=": "仙侠",
        "https://www.qidian.com/rank/readIndex?style=1&chn=4&page=": "都市",
        "https://www.qidian.com/rank/readIndex?style=1&chn=15&page=": "现实",
        "https://www.qidian.com/rank/readIndex?style=1&chn=6&page=": "军事",
        "https://www.qidian.com/rank/readIndex?style=1&chn=5&page=": "历史",
        "https://www.qidian.com/rank/readIndex?style=1&chn=7&page=": "游戏",
        "https://www.qidian.com/rank/readIndex?style=1&chn=8&page=": "体育",
        "https://www.qidian.com/rank/readIndex?style=1&chn=9&page=": "科幻",
        "https://www.qidian.com/rank/readIndex?style=1&chn=10&page=": "悬疑",
        "https://www.qidian.com/rank/readIndex?style=1&chn=12&page=": "轻小说"
    }
    # 循环list_的索引既分类的ｕｒｌ进行遍历
    for nov in list_.keys():
        #1~26为小说排行榜的页数，每一类有２５页的小说，一页小说有２０本书
        for num in range(1, 26):
            #获得小说类型
            class_nov = list_[nov]
            #进入小说的分类
            xh = requests.get(url=nov + str(num), headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:60.0) Gecko/20100101 Firefox/60.0'})
            #代码美化，获取标签定位
            be = BeautifulSoup(xh.text, 'html.parser')
            #遍历２０本小说的ｕｒｌ，进入小说的详细页面获取信息
            for info in be.find_all(class_='book-mid-info'):
                # xhlist += "https"+info.h4.a.get('href')
                #如果获取信息成功返回１
                is_ok = Getinfo(str("https:" + info.h4.a.get('href')), class_nov, all_num=all_num)
                #如果成功ｅｘｃｌｅ的行数＋１
                if is_ok == 1:
                    all_num += 1


def save_excle(noveltitle, novelwriter, novelnum, novellike_all, novellike_week, class_nov, all_num):
    ws = wb['Sheet1']
    # 小说名称
    ws['A' + str(all_num)].value = noveltitle
    # 小说作者
    ws['B' + str(all_num)].value = novelwriter
    # 小说分类
    ws['C' + str(all_num)].value = class_nov

    # 小说的字数
    if '.' in novelnum:
        if novelnum[-3] == '.':
            ws['D' + str(all_num)].value = str(novelnum).replace('.', '') + '00'
        else:
            ws['D' + str(all_num)].value = str(novelnum).replace('.', '') + '000'
    else:
        ws['D' + str(all_num)].value = novelnum

    # 小说的总推荐数
    if '.' in novellike_all:
        if novellike_all[-3] == '.':
            ws['E' + str(all_num)].value = str(novellike_all).replace('.', '') + '00'
        else:
            ws['E' + str(all_num)].value = str(novellike_all).replace('.', '') + '000'
    else:
        ws['E' + str(all_num)].value = novellike_all

    # 小说的周推荐数
    if '.' in novellike_week:
        if novellike_week[-3] == '.':
            ws['F' + str(all_num)].value = str(novellike_week).replace('.', '') + '00'
        else:
            ws['F' + str(all_num)].value = str(novellike_week).replace('.', '') + '000'
    else:
        ws['F' + str(all_num)].value = novellike_week

    # 保存修改后的ｅｘｃｌｅ文件
    wb.save('zhanghan.xlsx')


if __name__ == '__main__':
    # 创建操作ｅｘｃｌｅ的对象
    wb = openpyxl.load_workbook('zhanghan.xlsx')
    # 开始爬取，从小说的分类列表开始爬取，小说列表作为主目录，小说的排行页为二级目录，小说为最低级，层层递归爬取
    Getnovellist()
